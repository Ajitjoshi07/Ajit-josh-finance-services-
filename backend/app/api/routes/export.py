"""
Export Routes — Excel downloads with token-based auth
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from typing import Optional
import io
from decimal import Decimal
from datetime import datetime

from app.db.database import get_db
from app.models.models import (
    User, ClientProfile, Transaction, Document,
    TDSRecord, JournalEntry
)
from app.core.security import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])


async def get_user_from_token(token: Optional[str], db: AsyncSession) -> User:
    """Authenticate via query param token"""
    if not token:
        raise HTTPException(401, "Not authenticated — include token parameter")
    try:
        from app.core.security import decode_token
        payload = decode_token(token)
        user_id = payload.get("sub")
        result = await db.execute(select(User).where(User.id == int(user_id)))
        user = result.scalar_one_or_none()
        if not user:
            raise HTTPException(401, "User not found")
        return user
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(401, "Invalid token")


async def resolve_profile(current_user: User, db: AsyncSession, client_id: Optional[int]) -> ClientProfile:
    if current_user.role == "client":
        result = await db.execute(
            select(ClientProfile).where(ClientProfile.user_id == current_user.id)
        )
    else:
        if client_id:
            result = await db.execute(
                select(ClientProfile).where(ClientProfile.id == client_id)
            )
        else:
            # Admin without client_id: pick first active client
            result = await db.execute(
                select(ClientProfile).where(ClientProfile.is_active == True).limit(1)
            )
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "No client profile found. Please select a client or create one first.")
    return p


def make_workbook():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        return openpyxl, Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not available")


def style_header(ws, row, headers, openpyxl, Font, PatternFill):
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = font
        cell.fill = fill


@router.get("/excel/transactions")
async def export_transactions(
    financial_year: str = "2024-25",
    client_id: Optional[int] = None,
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    current_user = await get_user_from_token(token, db)
    profile = await resolve_profile(current_user, db, client_id)
    openpyxl, Font, PatternFill, Alignment = make_workbook()

    result = await db.execute(
        select(Transaction).where(
            and_(Transaction.client_id == profile.id,
                 Transaction.financial_year == financial_year)
        ).order_by(Transaction.invoice_date)
    )
    txns = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Register"
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Transaction Register — {profile.business_name} — FY {financial_year}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"PAN: {profile.pan or '—'} | GSTIN: {profile.gstin or 'Not Registered'}"

    headers = ["#", "Type", "Invoice No", "Date", "Party Name", "GSTIN",
               "Taxable Amt", "CGST", "SGST", "IGST", "Total Amt", "TDS", "Month", "Validated"]
    style_header(ws, 4, headers, openpyxl, Font, PatternFill)

    for i, t in enumerate(txns, 1):
        ws.append([
            i, t.transaction_type, t.invoice_number or "",
            str(t.invoice_date) if t.invoice_date else "",
            t.party_name or "", t.party_gstin or "",
            float(t.taxable_amount or 0), float(t.cgst_amount or 0),
            float(t.sgst_amount or 0), float(t.igst_amount or 0),
            float(t.total_amount or 0), float(t.tds_amount or 0),
            t.month or "", "Yes" if t.is_validated else "Pending"
        ])

    # Total row
    if txns:
        ws.append(["", "TOTAL", "", "", "", "",
                   sum(float(t.taxable_amount or 0) for t in txns),
                   sum(float(t.cgst_amount or 0) for t in txns),
                   sum(float(t.sgst_amount or 0) for t in txns),
                   sum(float(t.igst_amount or 0) for t in txns),
                   sum(float(t.total_amount or 0) for t in txns),
                   sum(float(t.tds_amount or 0) for t in txns), "", ""])
        for c in range(1, 15):
            ws.cell(ws.max_row, c).font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4, 30
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{profile.business_name}_{financial_year}_Transactions.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/excel/gst")
async def export_gst(
    financial_year: str = "2024-25",
    client_id: Optional[int] = None,
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    current_user = await get_user_from_token(token, db)
    profile = await resolve_profile(current_user, db, client_id)
    openpyxl, Font, PatternFill, Alignment = make_workbook()

    from app.services.tax.gst_engine import GSTEngine
    engine = GSTEngine(db)
    summary = await engine.get_monthly_summary(profile.id, financial_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Summary"
    ws.merge_cells("A1:H1")
    ws["A1"] = f"GST Summary — {profile.business_name} — FY {financial_year}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"GSTIN: {profile.gstin or 'Not Registered'}"

    headers = ["Month", "Year", "Total Sales (₹)", "Total Purchases (₹)",
               "Output GST (₹)", "Input ITC (₹)", "Net GST Payable (₹)", "Status"]
    style_header(ws, 4, headers, openpyxl, Font, PatternFill)

    total_payable = 0
    for m in summary:
        net = float(m["net_gst_payable"])
        total_payable += net
        ws.append([
            m.get("month_name", m["month"]), m["year"],
            float(m["total_sales"]), float(m["total_purchases"]),
            float(m["output_gst"]), float(m["input_gst"]),
            net, m["filing_status"].upper()
        ])

    ws.append(["ANNUAL TOTAL", "", "", "", "", "", total_payable, ""])
    for c in range(1, 9):
        ws.cell(ws.max_row, c).font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{profile.business_name}_{financial_year}_GST_Summary.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/excel/tds")
async def export_tds(
    financial_year: str = "2024-25",
    client_id: Optional[int] = None,
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    current_user = await get_user_from_token(token, db)
    profile = await resolve_profile(current_user, db, client_id)
    openpyxl, Font, PatternFill, Alignment = make_workbook()

    result = await db.execute(
        select(TDSRecord).where(
            and_(TDSRecord.client_id == profile.id, TDSRecord.financial_year == financial_year)
        ).order_by(TDSRecord.quarter)
    )
    records = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDS Records"
    ws["A1"] = f"TDS Records — {profile.business_name} — FY {financial_year}"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["Quarter", "Section", "Deductee Name", "Deductee PAN",
               "Payment Date", "Payment Amt (₹)", "TDS Rate %", "TDS Amount (₹)", "Deposited", "Challan No"]
    style_header(ws, 3, headers, openpyxl, Font, PatternFill)

    for r in records:
        ws.append([
            f"Q{r.quarter}", r.section or "", r.deductee_name or "", r.deductee_pan or "",
            str(r.payment_date) if r.payment_date else "",
            float(r.payment_amount or 0), r.tds_rate or 0,
            float(r.tds_amount or 0), "Yes" if r.deposited else "No",
            r.challan_number or ""
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{profile.business_name}_{financial_year}_TDS.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/excel/complete")
async def export_complete(
    financial_year: str = "2024-25",
    client_id: Optional[int] = None,
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Complete CA file — ICAI format multi-sheet Excel"""
    current_user = await get_user_from_token(token, db)
    profile = await resolve_profile(current_user, db, client_id)
    openpyxl, Font, PatternFill, Alignment = make_workbook()
    from openpyxl.styles import Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    DARK_BLUE = "1E3A5F"
    LIGHT_BLUE = "D6E4F0"
    GOLD = "C9A84C"
    WHITE = "FFFFFF"
    GRAY = "F5F5F5"

    def hdr_font(size=11, bold=True, color=WHITE):
        return Font(bold=bold, size=size, color=color)

    def hdr_fill(color=DARK_BLUE):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def write_icai_header(ws, title, subtitle, business_name, fy):
        ws.merge_cells("A1:H1")
        ws["A1"] = "AJIT JOSHI FINANCE SERVICES"
        ws["A1"].font = Font(bold=True, size=14, color=DARK_BLUE)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:H2")
        ws["A2"] = title
        ws["A2"].font = Font(bold=True, size=13, color=DARK_BLUE)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:H3")
        ws["A3"] = f"{business_name}   |   {subtitle}   |   FY {fy}"
        ws["A3"].font = Font(size=10, italic=True, color="555555")
        ws["A3"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 14

    def style_header_row(ws, row, headers, widths=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = hdr_font()
            c.fill = hdr_fill()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
            if widths and col <= len(widths):
                ws.column_dimensions[get_column_letter(col)].width = widths[col-1]
        ws.row_dimensions[row].height = 22

    def fmt_amt(v):
        return float(v or 0)

    def amt_cell(ws, row, col, val, bold=False, color=None):
        c = ws.cell(row=row, column=col, value=fmt_amt(val))
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal="right")
        c.border = thin_border()
        if bold:
            c.font = Font(bold=True, color=color or "000000")
        return c

    def lbl_cell(ws, row, col, val, bold=False, indent=0, bg=None):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left", indent=indent)
        if bold:
            c.font = Font(bold=True)
        if bg:
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        return c

    # ── Sheet 1: Cover ──────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Cover Page"
    ws_cover.sheet_view.showGridLines = False
    ws_cover.merge_cells("B2:F2")
    ws_cover["B2"] = "CLIENT FINANCIAL FILE"
    ws_cover["B2"].font = Font(bold=True, size=20, color=DARK_BLUE)
    ws_cover["B2"].alignment = Alignment(horizontal="center")
    ws_cover.row_dimensions[2].height = 35
    info = [
        ("CA Firm", "Ajit Joshi Finance Services"),
        ("Prepared Under", "ICAI Accounting Standards"),
        ("", ""),
        ("Client Business Name", profile.business_name or "—"),
        ("PAN", profile.pan or "—"),
        ("GSTIN", profile.gstin or "Not Registered"),
        ("Business Type", profile.business_type or "—"),
        ("State", profile.state or "—"),
        ("Financial Year", financial_year),
        ("Assessment Year", f"AY 20{financial_year.split('-')[1]}-{int(financial_year.split('-')[1]) + 1:02d}"),
        ("", ""),
        ("Date of Preparation", datetime.now().strftime("%d-%m-%Y")),
        ("", ""),
        ("DISCLAIMER", "This document is prepared for CA review. All figures subject to verification."),
    ]
    for i, (k, v) in enumerate(info, 4):
        c = ws_cover.cell(i, 2, k)
        c.font = Font(bold=True, size=11)
        ws_cover.cell(i, 3, ":")
        c2 = ws_cover.cell(i, 4, v)
        if k == "DISCLAIMER":
            c2.font = Font(italic=True, size=10, color="888888")
        ws_cover.row_dimensions[i].height = 18
    ws_cover.column_dimensions["B"].width = 28
    ws_cover.column_dimensions["C"].width = 3
    ws_cover.column_dimensions["D"].width = 45

    # ── Sheet 2: Transaction Register ──────────────────────────────
    txn_result = await db.execute(
        select(Transaction).where(
            and_(Transaction.client_id == profile.id, Transaction.financial_year == financial_year)
        ).order_by(Transaction.invoice_date)
    )
    txns = txn_result.scalars().all()
    ws_txn = wb.create_sheet("Transaction Register")
    write_icai_header(ws_txn, "TRANSACTION REGISTER", "All Entries", profile.business_name, financial_year)
    hdrs = ["#", "Type", "Invoice No.", "Date", "Party Name", "Party GSTIN", "Taxable Amt", "CGST", "SGST", "IGST", "Total Amt", "TDS", "Month", "Status"]
    style_header_row(ws_txn, 5, hdrs, [4, 14, 14, 12, 22, 18, 14, 10, 10, 10, 14, 10, 8, 10])
    for i, t in enumerate(txns, 1):
        r = 5 + i
        ws_txn.cell(r, 1, i).border = thin_border()
        ws_txn.cell(r, 2, (t.transaction_type or "").title()).border = thin_border()
        ws_txn.cell(r, 3, t.invoice_number or "").border = thin_border()
        ws_txn.cell(r, 4, str(t.invoice_date or "")).border = thin_border()
        ws_txn.cell(r, 5, t.party_name or "").border = thin_border()
        ws_txn.cell(r, 6, t.party_gstin or "").border = thin_border()
        amt_cell(ws_txn, r, 7, t.taxable_amount)
        amt_cell(ws_txn, r, 8, t.cgst_amount)
        amt_cell(ws_txn, r, 9, t.sgst_amount)
        amt_cell(ws_txn, r, 10, t.igst_amount)
        amt_cell(ws_txn, r, 11, t.total_amount, bold=True)
        amt_cell(ws_txn, r, 12, t.tds_amount)
        ws_txn.cell(r, 13, t.month or "").border = thin_border()
        status_c = ws_txn.cell(r, 14, "Approved" if t.is_validated else "Pending")
        status_c.border = thin_border()
        status_c.font = Font(color="007B3A" if t.is_validated else "CC5500")
    tot_r = 6 + len(txns)
    for col in range(1, 15):
        c = ws_txn.cell(tot_r, col, "TOTAL" if col == 1 else "")
        c.font = Font(bold=True); c.fill = hdr_fill(LIGHT_BLUE); c.border = thin_border()
    for col, field in [(7, "taxable_amount"), (8, "cgst_amount"), (9, "sgst_amount"), (10, "igst_amount"), (11, "total_amount"), (12, "tds_amount")]:
        amt_cell(ws_txn, tot_r, col, sum(float(getattr(t, field) or 0) for t in txns), bold=True)
    ws_txn.freeze_panes = "A6"

    # ── Sheet 3: GST Summary ICAI ───────────────────────────────────
    from app.services.tax.gst_engine import GSTEngine
    gst_engine = GSTEngine(db)
    gst_summary = await gst_engine.get_monthly_summary(profile.id, financial_year)
    ws_gst = wb.create_sheet("GST Summary")
    write_icai_header(ws_gst, "GST MONTHLY SUMMARY (GSTR-3B)", f"GSTIN: {profile.gstin or 'Not Registered'}", profile.business_name, financial_year)
    gst_hdrs = ["Month", "Year", "Taxable Sales", "Output CGST", "Output SGST", "Output IGST", "Total Output GST", "Input ITC", "Net GST Payable", "GSTR-1 Status", "GSTR-3B Status"]
    style_header_row(ws_gst, 5, gst_hdrs, [12, 8, 16, 14, 14, 14, 18, 14, 16, 14, 14])
    tot_payable = 0
    for i, m in enumerate(gst_summary, 1):
        r = 5 + i
        ws_gst.cell(r, 1, m.get("month_name", m["month"])).border = thin_border()
        ws_gst.cell(r, 2, m["year"]).border = thin_border()
        amt_cell(ws_gst, r, 3, m["total_sales"])
        out_gst = float(m["output_gst"])
        amt_cell(ws_gst, r, 4, out_gst / 2)
        amt_cell(ws_gst, r, 5, out_gst / 2)
        amt_cell(ws_gst, r, 6, 0)
        amt_cell(ws_gst, r, 7, out_gst, bold=True)
        amt_cell(ws_gst, r, 8, m["input_gst"])
        net = float(m["net_gst_payable"])
        tot_payable += net
        c = amt_cell(ws_gst, r, 9, net, bold=True, color="8B0000" if net > 0 else "006400")
        status = m["filing_status"].upper()
        for col in [10, 11]:
            sc = ws_gst.cell(r, col, status)
            sc.border = thin_border()
            sc.alignment = Alignment(horizontal="center")
            sc.font = Font(color="007B3A" if status == "FILED" else "CC5500", bold=True)
    tot_r = 6 + len(gst_summary)
    for col in range(1, 12):
        c = ws_gst.cell(tot_r, col, "ANNUAL TOTAL" if col == 1 else "")
        c.font = Font(bold=True); c.fill = hdr_fill(DARK_BLUE); c.font = Font(bold=True, color=WHITE)
        c.border = thin_border()
    amt_cell(ws_gst, tot_r, 9, tot_payable, bold=True, color=WHITE)
    ws_gst.cell(tot_r, 9).fill = hdr_fill(DARK_BLUE)
    ws_gst.freeze_panes = "A6"

    # ── Sheet 4: TDS Records ────────────────────────────────────────
    tds_result = await db.execute(
        select(TDSRecord).where(
            and_(TDSRecord.client_id == profile.id, TDSRecord.financial_year == financial_year)
        ).order_by(TDSRecord.quarter, TDSRecord.section)
    )
    tds_records = tds_result.scalars().all()
    ws_tds = wb.create_sheet("TDS Records")
    write_icai_header(ws_tds, "TDS DEDUCTION REGISTER", "As per Income Tax Act 1961", profile.business_name, financial_year)
    tds_hdrs = ["Quarter", "Section", "Nature of Payment", "Deductee Name", "Deductee PAN", "Payment Date", "Payment Amount", "TDS Rate %", "TDS Amount", "Deposited", "Challan No."]
    style_header_row(ws_tds, 5, tds_hdrs, [10, 10, 22, 22, 14, 12, 16, 12, 14, 10, 14])
    q_totals = {}
    for i, r in enumerate(tds_records, 1):
        row = 5 + i
        ws_tds.cell(row, 1, f"Q{r.quarter}").border = thin_border()
        ws_tds.cell(row, 2, r.section or "").border = thin_border()
        nature = {"194C": "Contractor", "194J": "Professional/Technical", "194H": "Commission", "194I": "Rent", "194A": "Interest", "192": "Salary"}.get(r.section or "", "Other")
        ws_tds.cell(row, 3, nature).border = thin_border()
        ws_tds.cell(row, 4, r.deductee_name or "").border = thin_border()
        ws_tds.cell(row, 5, r.deductee_pan or "").border = thin_border()
        ws_tds.cell(row, 6, str(r.payment_date or "")).border = thin_border()
        amt_cell(ws_tds, row, 7, r.payment_amount)
        ws_tds.cell(row, 8, float(r.tds_rate or 0)).border = thin_border()
        amt_cell(ws_tds, row, 9, r.tds_amount, bold=True)
        dep = ws_tds.cell(row, 10, "Yes" if r.deposited else "No")
        dep.border = thin_border(); dep.font = Font(color="007B3A" if r.deposited else "CC0000")
        ws_tds.cell(row, 11, r.challan_number or "").border = thin_border()
    tot_r = 6 + len(tds_records)
    for col in range(1, 12):
        c = ws_tds.cell(tot_r, col, "TOTAL" if col == 1 else "")
        c.font = Font(bold=True, color=WHITE); c.fill = hdr_fill(); c.border = thin_border()
    amt_cell(ws_tds, tot_r, 7, sum(float(r.payment_amount or 0) for r in tds_records), bold=True)
    amt_cell(ws_tds, tot_r, 9, sum(float(r.tds_amount or 0) for r in tds_records), bold=True)
    ws_tds.freeze_panes = "A6"

    # ── Sheet 5: Manufacturing Account (ICAI) ──────────────────────
    from app.services.reports.financial_statements import FinancialStatementsService
    svc = FinancialStatementsService(db)
    pl = await svc.get_profit_loss(profile.id, financial_year)
    inc = pl.get("income", {})
    cogs = pl.get("cost_of_goods", {})
    exp = pl.get("expenses", {})
    safe = lambda v: float(v or 0)
    gross_profit = float(pl.get("gross_profit", 0))
    net_profit = float(pl.get("net_profit", 0))
    raw_mat = safe(cogs.get("Raw Material Consumed")) or safe(cogs.get("Purchases", 0))
    dir_lab = safe(exp.get("Direct Labour", 0))
    fact_oh = safe(exp.get("Factory Overhead", 0))
    open_wip = safe(cogs.get("Opening WIP", 0))
    close_wip = safe(cogs.get("Closing WIP", 0))
    cost_prod = raw_mat + dir_lab + fact_oh + open_wip - close_wip
    open_stk = safe(cogs.get("Opening Stock", 0))
    close_stk = safe(cogs.get("Closing Stock", 0))
    total_sales = safe(inc.get("Sales Revenue", 0)) + safe(inc.get("Other Operating Income", 0))

    def icai_two_col_sheet(wb, sheet_name, title, subtitle, fy_year, rows):
        """rows = list of (dr_label, dr_amt, cr_label, cr_amt, is_heading, is_total)"""
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:F1")
        ws["A1"] = "AJIT JOSHI FINANCE SERVICES"
        ws["A1"].font = Font(bold=True, size=12, color=DARK_BLUE)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F2")
        ws["A2"] = title
        ws["A2"].font = Font(bold=True, size=13, color=DARK_BLUE)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:F3")
        ws["A3"] = subtitle
        ws["A3"].font = Font(size=10, italic=True, color="555555")
        ws["A3"].alignment = Alignment(horizontal="center")
        # Column headers
        ws.merge_cells("A4:B4")
        ws["A4"] = "Dr — Particulars"
        ws["A4"].font = hdr_font()
        ws["A4"].fill = hdr_fill()
        ws["A4"].alignment = Alignment(horizontal="center")
        ws["C4"] = "Amount (Rs.)"
        ws["C4"].font = hdr_font()
        ws["C4"].fill = hdr_fill()
        ws["C4"].alignment = Alignment(horizontal="center")
        ws.merge_cells("D4:E4")
        ws["D4"] = "Cr — Particulars"
        ws["D4"].font = hdr_font()
        ws["D4"].fill = hdr_fill()
        ws["D4"].alignment = Alignment(horizontal="center")
        ws["F4"] = "Amount (Rs.)"
        ws["F4"].font = hdr_font()
        ws["F4"].fill = hdr_fill()
        ws["F4"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[4].height = 22
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 4
        ws.column_dimensions["E"].width = 32
        ws.column_dimensions["F"].width = 18
        for i, (dl, da, cl, ca, is_hd, is_tot) in enumerate(rows, 5):
            bg = DARK_BLUE if is_hd else (LIGHT_BLUE if is_tot else None)
            txt_color = WHITE if is_hd else "000000"
            for col in range(1, 7):
                c = ws.cell(i, col)
                c.border = thin_border()
                if bg:
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            if is_hd:
                ws.merge_cells(f"A{i}:F{i}")
                ws.cell(i, 1, dl).font = Font(bold=True, color=WHITE, size=10)
                ws.cell(i, 1).fill = hdr_fill(DARK_BLUE)
            else:
                ws.cell(i, 2, dl).font = Font(bold=is_tot, color=txt_color)
                ws.cell(i, 2).alignment = Alignment(indent=1)
                if da is not None:
                    c = ws.cell(i, 3, da)
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(bold=is_tot, color=txt_color)
                ws.cell(i, 5, cl).font = Font(bold=is_tot, color=txt_color)
                ws.cell(i, 5).alignment = Alignment(indent=1)
                if ca is not None:
                    c = ws.cell(i, 6, ca)
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")
                    c.font = Font(bold=is_tot, color=txt_color)
        return ws

    fy_year_str = f"20{financial_year.split('-')[1]}" if len(financial_year.split('-')[1]) == 2 else financial_year.split('-')[1]

    # Manufacturing Account rows
    mfg_rows = [
        ("Opening Stock of Raw Material", safe(cogs.get("Opening Stock of Raw Material", 0)), "Closing Stock of Raw Material", safe(cogs.get("Closing Stock of Raw Material", 0)), False, False),
        ("Raw Material Purchases", safe(cogs.get("Raw Material Purchases", 0)), "Closing WIP", close_wip, False, False),
        ("Less: Purchase Returns", safe(cogs.get("Purchase Returns", 0)), "", None, False, False),
        ("Raw Material Consumed", raw_mat, "", None, False, True),
        ("Opening WIP", open_wip, "", None, False, False),
        ("Direct Labour / Wages", dir_lab, "", None, False, False),
        ("Power & Fuel (Factory)", safe(exp.get("Power & Fuel", 0)), "", None, False, False),
        ("Factory Overhead", fact_oh, "", None, False, False),
        ("Packing Materials", safe(exp.get("Packing Materials", 0)), "", None, False, False),
        ("Repairs & Maintenance (Factory)", safe(exp.get("Repairs & Maintenance", 0)), "", None, False, False),
        ("Depreciation on Factory Assets", safe(exp.get("Depreciation (Factory)", 0)), "", None, False, False),
        ("Other Manufacturing Expenses", safe(exp.get("Other Manufacturing Expenses", 0)), "", None, False, False),
        ("Cost of Production c/d", cost_prod, "Cost of Production b/d", cost_prod, False, True),
    ]
    icai_two_col_sheet(wb, "Manufacturing Account", "MANUFACTURING ACCOUNT", f"For the year ended 31st March, {fy_year_str}", fy_year_str, mfg_rows)

    # Trading Account rows
    total_dr_trading = open_stk + cost_prod + safe(cogs.get("Purchases", 0))
    trading_rows = [
        ("Opening Stock of Finished Goods", open_stk, "Sales (Gross)", safe(inc.get("Sales Revenue", 0)), False, False),
        ("Cost of Production b/d", cost_prod, "Less: Sales Returns", safe(cogs.get("Sales Returns", 0)), False, False),
        ("Add: Purchases (Trading Goods)", safe(cogs.get("Purchases", 0)), "Net Sales", total_sales, False, True),
        ("Less: Purchase Returns", safe(cogs.get("Purchase Returns", 0)), "Other Operating Income", safe(inc.get("Other Operating Income", 0)), False, False),
        ("Carriage Outward", safe(exp.get("Carriage Outward", 0)), "Closing Stock of Finished Goods", close_stk, False, False),
        ("Custom Duty / Octroi", safe(exp.get("Custom Duty", 0)), "", None, False, False),
        ("Gross Profit c/d" if gross_profit >= 0 else "", gross_profit if gross_profit >= 0 else None, "Gross Loss c/d" if gross_profit < 0 else "", abs(gross_profit) if gross_profit < 0 else None, False, True),
    ]
    icai_two_col_sheet(wb, "Trading Account", "TRADING ACCOUNT", f"For the year ended 31st March, {fy_year_str}", fy_year_str, trading_rows)

    # P&L rows
    total_exp = sum(safe(v) for v in exp.values())
    total_income_other = sum(safe(v) for v in inc.values())
    pl_rows = [
        ("" if gross_profit >= 0 else "Gross Loss b/d", None if gross_profit >= 0 else abs(gross_profit), "Gross Profit b/d" if gross_profit >= 0 else "", gross_profit if gross_profit >= 0 else None, False, False),
        ("Salaries & Wages (Office/Admin)", safe(exp.get("Salaries", 0)), "Commission Received", safe(inc.get("Commission Income", 0)), False, False),
        ("Rent, Rates & Taxes", safe(exp.get("Rent", 0)), "Discount Received", safe(inc.get("Discount Received", 0)), False, False),
        ("Electricity & Water (Office)", safe(exp.get("Utilities", 0)), "Interest / Dividend Received", safe(inc.get("Interest Income", 0)) + safe(inc.get("Dividend Income", 0)), False, False),
        ("Printing & Stationery", safe(exp.get("Printing & Stationery", 0)), "Rental Income", safe(inc.get("Rental Income", 0)), False, False),
        ("Postage & Courier", safe(exp.get("Postage", 0)), "Profit on Sale of Fixed Asset", safe(inc.get("Profit on Sale of Assets", 0)), False, False),
        ("Telephone & Internet", safe(exp.get("Telephone", 0)), "Miscellaneous Income", safe(inc.get("Other Income", 0)), False, False),
        ("Travelling & Conveyance", safe(exp.get("Travelling", 0)), "", None, False, False),
        ("Advertisement & Marketing", safe(exp.get("Advertisement", 0)), "", None, False, False),
        ("Professional / Legal Fees", safe(exp.get("Professional Fees", 0)), "", None, False, False),
        ("Audit Fees", safe(exp.get("Audit Fees", 0)), "", None, False, False),
        ("Bank Charges & Interest on OD", safe(exp.get("Bank Charges", 0)) + safe(exp.get("Interest on OD", 0)), "", None, False, False),
        ("Depreciation (Non-factory)", safe(exp.get("Depreciation", 0)), "", None, False, False),
        ("Bad Debts / Provision", safe(exp.get("Bad Debts", 0)) + safe(exp.get("Provision Bad Debts", 0)), "", None, False, False),
        ("Discount Allowed", safe(exp.get("Discount Allowed", 0)), "", None, False, False),
        ("Insurance Premium", safe(exp.get("Insurance", 0)), "", None, False, False),
        ("Staff Welfare", safe(exp.get("Staff Welfare", 0)), "", None, False, False),
        ("GST Late Fees & Penalty", safe(exp.get("GST Late Fees", 0)), "", None, False, False),
        ("Miscellaneous Expenses", safe(exp.get("Miscellaneous", 0)), "", None, False, False),
        ("Net Profit transferred to Capital" if net_profit >= 0 else "", net_profit if net_profit >= 0 else None, "Net Loss transferred to Capital" if net_profit < 0 else "", abs(net_profit) if net_profit < 0 else None, False, True),
        ("TOTAL", total_exp + (0 if gross_profit >= 0 else abs(gross_profit)), "TOTAL", total_income_other + (gross_profit if gross_profit >= 0 else 0), False, True),
    ]
    icai_two_col_sheet(wb, "Profit & Loss Account", "PROFIT & LOSS ACCOUNT", f"For the year ended 31st March, {fy_year_str}", fy_year_str, pl_rows)

    # Balance Sheet
    bs = await svc.get_balance_sheet(profile.id, financial_year)
    safe_bs = lambda path, key: float((bs.get(path) or {}).get(key) or 0)
    bs_rows = [
        ("CAPITAL & RESERVES", None, "FIXED ASSETS", None, True, False),
        ("Capital Account (Opening)", safe_bs("capital", "Owner Capital"), "Land & Building", safe_bs("assets", "Land & Building"), False, False),
        ("Add: Net Profit for the Year", net_profit if net_profit > 0 else 0, "Plant & Machinery", safe_bs("assets", "Plant & Machinery"), False, False),
        ("Less: Drawings", safe_bs("capital", "Drawings"), "Furniture & Fixtures", safe_bs("assets", "Furniture"), False, False),
        ("Reserves & Surplus", safe_bs("capital", "Reserves"), "Vehicles", safe_bs("assets", "Vehicles"), False, False),
        ("Share Capital", safe_bs("capital", "Share Capital"), "Less: Accumulated Depreciation", safe_bs("assets", "Accumulated Depreciation"), False, False),
        ("LONG-TERM LIABILITIES", None, "Net Fixed Assets", float(bs.get("assets", {}).get("total_fixed") or 0), True, False),
        ("Secured Term Loans", safe_bs("liabilities", "Secured Loans"), "CWIP", safe_bs("assets", "CWIP"), False, False),
        ("Unsecured Loans", safe_bs("liabilities", "Unsecured Loans"), "Intangible Assets", safe_bs("assets", "Intangibles"), False, False),
        ("Debentures", safe_bs("liabilities", "Debentures"), "Long-term Investments", 0, False, False),
        ("CURRENT LIABILITIES", None, "CURRENT ASSETS", None, True, False),
        ("Sundry Creditors", safe_bs("liabilities", "Accounts Payable"), "Closing Stock", close_stk, False, False),
        ("Bills Payable", safe_bs("liabilities", "Bills Payable"), "Sundry Debtors", safe_bs("assets", "Accounts Receivable"), False, False),
        ("Bank Overdraft / CC", safe_bs("liabilities", "Bank OD"), "Bills Receivable", safe_bs("assets", "Bills Receivable"), False, False),
        ("Advance from Customers", safe_bs("liabilities", "Advance from Customers"), "Cash in Hand", safe_bs("assets", "Cash"), False, False),
        ("Outstanding Expenses", safe_bs("liabilities", "Outstanding Expenses"), "Bank Balance", safe_bs("assets", "Bank Account"), False, False),
        ("GST Payable (CGST+SGST+IGST)", safe_bs("liabilities", "GST Payable"), "Prepaid Expenses", safe_bs("assets", "Prepaid Expenses"), False, False),
        ("TDS Payable", safe_bs("liabilities", "TDS Payable"), "Advance Tax / TDS Receivable", safe_bs("assets", "TDS Receivable"), False, False),
        ("Salaries Payable", safe_bs("liabilities", "Salary Payable"), "Accrued Income", safe_bs("assets", "Accrued Income"), False, False),
        ("Income Tax / Provision", safe_bs("liabilities", "Tax Payable"), "Short-term Investments", safe_bs("assets", "Short Term Investments"), False, False),
        ("TOTAL LIABILITIES & CAPITAL", float(bs.get("total_liabilities_capital") or 0), "TOTAL ASSETS", float((bs.get("assets") or {}).get("total_assets") or 0), False, True),
    ]
    icai_two_col_sheet(wb, "Balance Sheet", "BALANCE SHEET", f"As at 31st March, {fy_year_str}", fy_year_str, bs_rows)

    # ITR Summary sheet
    from app.services.tax.itr_engine import ITREngine
    itr_engine = ITREngine(db)
    itr = await itr_engine.compute_itr(profile.id, financial_year)
    ws_itr = wb.create_sheet("ITR Summary")
    ws_itr.sheet_view.showGridLines = False
    write_icai_header(ws_itr, "ITR COMPUTATION SHEET", f"AY 20{financial_year.split('-')[1]}-{int(financial_year.split('-')[1])+1:02d} | New Tax Regime Sec 115BAC", profile.business_name, financial_year)
    itr_rows = [
        ("PAN", profile.pan or "—"),
        ("ITR Form Type", "ITR-4 Sugam (Presumptive Business Income u/s 44AD/44ADA)"),
        ("", ""),
        ("Gross Total Income (GTI)", f"Rs. {float(itr.gross_income):,.2f}"),
        ("Less: Total Deductions (Chap VI-A)", f"Rs. {float(itr.total_deductions):,.2f}"),
        ("Net Taxable Income", f"Rs. {float(itr.taxable_income):,.2f}"),
        ("", ""),
        ("Income Tax at Applicable Slabs", f"Rs. {float(itr.tax_liability):,.2f}"),
        ("Add: Health & Education Cess @4%", f"Rs. {float(itr.tax_liability) * 0.04:,.2f}"),
        ("Total Tax After Cess", f"Rs. {float(itr.tax_liability) * 1.04:,.2f}"),
        ("", ""),
        ("Less: TDS Deducted at Source", f"Rs. {float(itr.tds_paid):,.2f}"),
        ("Less: Advance Tax Paid", "Rs. 0.00"),
        ("Net Tax Payable / (Refund)", f"Rs. {float(itr.net_tax_payable):,.2f}"),
    ]
    for i, (k, v) in enumerate(itr_rows, 5):
        c1 = ws_itr.cell(i, 2, k)
        c2 = ws_itr.cell(i, 4, v)
        c1.border = thin_border(); c2.border = thin_border()
        if k:
            c1.font = Font(bold=True, size=11)
            c2.font = Font(size=11)
        ws_itr.row_dimensions[i].height = 18
    ws_itr.column_dimensions["B"].width = 38
    ws_itr.column_dimensions["C"].width = 4
    ws_itr.column_dimensions["D"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{profile.business_name}_{financial_year}_Complete_CA_File_ICAI.xlsx".replace(" ", "_")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/ca-file-summary")
async def ca_file_summary(
    financial_year: str = "2024-25",
    client_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    profile = await resolve_profile(current_user, db, client_id)

    txn_count = await db.execute(
        select(func.count(Transaction.id)).where(
            and_(Transaction.client_id == profile.id, Transaction.financial_year == financial_year)
        )
    )
    doc_count = await db.execute(
        select(func.count(Document.id)).where(
            and_(Document.client_id == profile.id, Document.financial_year == financial_year)
        )
    )
    sales_sum = await db.execute(
        select(func.coalesce(func.sum(Transaction.total_amount), 0)).where(
            and_(Transaction.client_id == profile.id, Transaction.financial_year == financial_year,
                 Transaction.transaction_type == "sales", Transaction.is_validated == True)
        )
    )
    purch_sum = await db.execute(
        select(func.coalesce(func.sum(Transaction.total_amount), 0)).where(
            and_(Transaction.client_id == profile.id, Transaction.financial_year == financial_year,
                 Transaction.transaction_type == "purchase", Transaction.is_validated == True)
        )
    )

    from app.services.tax.gst_engine import GSTEngine
    gst_engine = GSTEngine(db)
    gst_summary = await gst_engine.get_monthly_summary(profile.id, financial_year)
    total_gst = sum(float(m["net_gst_payable"]) for m in gst_summary)

    from app.services.tax.itr_engine import ITREngine
    itr_engine = ITREngine(db)
    itr = await itr_engine.compute_itr(profile.id, financial_year)

    return {
        "client": {"name": profile.business_name, "pan": profile.pan, "gstin": profile.gstin, "fy": financial_year},
        "data_summary": {
            "total_transactions": txn_count.scalar(),
            "total_documents": doc_count.scalar(),
            "total_sales": float(sales_sum.scalar() or 0),
            "total_purchases": float(purch_sum.scalar() or 0),
        },
        "gst": {"total_payable": total_gst, "months_filed": sum(1 for m in gst_summary if m["filing_status"] == "filed")},
        "itr": {
            "gross_income": float(itr.gross_income),
            "taxable_income": float(itr.taxable_income),
            "tax_liability": float(itr.tax_liability),
            "net_payable": float(itr.net_tax_payable),
        },
    }
